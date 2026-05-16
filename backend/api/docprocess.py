import json
import logging
import os
import re
import time
import uuid
from datetime import datetime

import openpyxl
import pandas as pd
from flask import Blueprint, jsonify, request, send_file
from flask_jwt_extended import get_jwt_identity, jwt_required
from sqlalchemy import text

from backend.api.auth import log_audit
from backend.database import SessionLocal
from backend.services.rag_service import answer_question_with_sources

logger = logging.getLogger(__name__)

docprocess_bp = Blueprint("docprocess", __name__)

UPLOADS_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "uploads")
RESULTS_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "downloads")
os.makedirs(UPLOADS_DIR, exist_ok=True)
os.makedirs(RESULTS_DIR, exist_ok=True)

MAX_FILE_SIZE_MB = 10
MAX_QUESTIONS = 500


def _serialize_sources(value):
    if not value:
        return []
    if isinstance(value, str):
        try:
            return json.loads(value)
        except Exception:
            return []
    return value


def _refresh_job_counters(db, batch_id):
    stats = db.execute(text("""
        SELECT
            COUNT(*) FILTER (WHERE status IN ('answered', 'accepted', 'error')) AS processed_count,
            COUNT(*) FILTER (WHERE status = 'accepted') AS accepted_count,
            COUNT(*) FILTER (WHERE status = 'error') AS error_count,
            COUNT(*) FILTER (WHERE cached = true) AS cached_count,
            COUNT(*) FILTER (WHERE status = 'processing') AS processing_count
        FROM batch_questions
        WHERE batch_id = :batch_id
    """), {"batch_id": batch_id}).mappings().fetchone()

    db.execute(text("""
        UPDATE batch_jobs
        SET processed_count = :processed_count,
            accepted_count = :accepted_count,
            error_count = :error_count,
            cached_count = :cached_count,
            updated_at = :now
        WHERE id = :batch_id
    """), {
        "batch_id": batch_id,
        "processed_count": stats["processed_count"] or 0,
        "accepted_count": stats["accepted_count"] or 0,
        "error_count": stats["error_count"] or 0,
        "cached_count": stats["cached_count"] or 0,
        "now": datetime.utcnow(),
    })
    return stats





_QUESTION_EXACT = {"question", "questions", "query", "queries", "q", "questionnaire", "questionnaires"}
_QUESTION_SUBSTR = ("question", "query", "questionnaire")


def _is_header_row(first_row_values):
    """Heuristic: if the first row has short labels (< 50 chars avg), it's a header.
    Long sentences mean it's probably data (headerless sheet)."""
    texts = [str(v).strip() for v in first_row_values if pd.notna(v) and str(v).strip()]
    if not texts:
        return False
    avg_len = sum(len(t) for t in texts) / len(texts)
    return avg_len < 50


def _match_question_column(columns):
    """Return the column name that matches question patterns, or None.
    Only considers reasonably short cells — a real column header is rarely
    longer than ~40 chars, and long paragraphs that happen to contain the
    word 'question' should not be treated as headers.
    """
    short_cols = [c for c in columns if len(str(c)) <= 40]
    for col in short_cols:
        normalized = re.sub(r'\s+', ' ', str(col)).strip().lower()
        if normalized in _QUESTION_EXACT:
            return col
    for col in short_cols:
        normalized = str(col).strip().lower()
        if any(t in normalized for t in _QUESTION_SUBSTR):
            return col
    return None


def _is_valid_question(text):
    """Return True only if text looks like a real question/sentence, not junk."""
    t = str(text).strip()
    if not t or t.lower() == "nan" or t.lower() == "none":
        return False
    # Too short to be a real question
    if len(t) < 8:
        return False
    # Must contain at least 3 word-like tokens (letters)
    words = re.findall(r'[a-zA-Z]{2,}', t)
    if len(words) < 3:
        return False
    # Ratio of letters to total chars — skip binary/code blobs
    alpha_ratio = sum(1 for c in t if c.isalpha()) / len(t)
    if alpha_ratio < 0.4:
        return False
    return True


def _dedupe_columns(cols):
    """Ensure column names are unique by appending suffixes to duplicates."""
    seen = {}
    out = []
    for c in cols:
        if c not in seen:
            seen[c] = 0
            out.append(c)
        else:
            seen[c] += 1
            out.append(f"{c}.{seen[c]}")
    return out


def _row_looks_like_header(row_vals):
    """A row 'looks like a header' if it has ≥2 short, mostly-textual cells.
    Excludes long sentences (data rows) and rows that are mostly numbers.
    Threshold ≤40 chars per cell matches typical column-header lengths
    (most are 5–25 chars; longer names like 'Follow-up  % Complete' are 22)."""
    vals = [str(v).strip() for v in row_vals if pd.notna(v) and str(v).strip()]
    if len(vals) < 2:
        return False
    if not all(len(v) <= 40 for v in vals):
        return False
    text_cells = sum(1 for v in vals if re.search(r'[A-Za-z]', v))
    # At least max(2, half-of-cells) must contain letters. Filters numeric
    # rows like ['Application Security', 13, 13, 0, 0, 0] (only 1 text cell).
    return text_cells >= max(2, len(vals) // 2)


def _find_header_row(raw_df, max_scan=15):
    """Find the first row that (a) looks like a header AND (b) contains a
    Question-like column. Returns the row index or None.

    The 'looks like a header' check is critical: without it, a data row
    that happens to contain the word 'question' (e.g. an index sheet row
    'General Questions | 0 | 11') would be mistaken for a header.
    """
    scan_limit = min(max_scan, raw_df.shape[0])
    for i in range(scan_limit):
        row = raw_df.iloc[i].tolist()
        if not _row_looks_like_header(row):
            continue
        row_vals = [str(v).strip() if pd.notna(v) else "" for v in row]
        if _match_question_column(row_vals) is not None:
            return i
    return None


def _find_likely_header_row(raw_df, max_scan=15):
    """Find a row that *looks* like a header but doesn't necessarily contain
    a Question column. Used to produce a clear error message ('found
    columns: [...]') when a data sheet is missing its Question column."""
    scan_limit = min(max_scan, raw_df.shape[0])
    for i in range(scan_limit):
        if _row_looks_like_header(raw_df.iloc[i].tolist()):
            return i
    return None


def _sheet_has_substantial_data(raw_df, min_cell_len=50, min_rows=1):
    """True if the sheet contains at least `min_rows` rows where some cell
    has ≥ `min_cell_len` characters of letter-bearing text.

    Used to distinguish 'real data sheet missing a Question column' (→ error)
    from 'utility / summary / index sheet' (→ skip silently).
    Threshold 50 chars matches typical real questionnaire content while
    excluding short index labels like 'Application Security' (20 chars).
    """
    long_rows = 0
    for i in range(min(200, raw_df.shape[0])):
        for v in raw_df.iloc[i].tolist():
            if pd.notna(v):
                s = str(v).strip()
                if len(s) >= min_cell_len and re.search(r'[A-Za-z]', s):
                    long_rows += 1
                    break
        if long_rows >= min_rows:
            return True
    return False


def _parse_sheet(sheet_name, raw_df):
    """Parse one sheet. Returns (questions_list, display_col_name, error_msg).
    raw_df is read with header=None so we can detect headerless sheets.

    Decision tree:
      1. Header row with a Question column found → parse rows below it.
      2. Likely header row found but no Question column:
           - If real data follows (long-text cells) → ERROR (user should rename column).
           - Otherwise → skip silently (Summary / index / instructions sheet).
      3. No header detectable but sheet contains long-text data → headerless
         fallback (pick the column with the longest average text).
      4. Otherwise → skip silently (empty / metadata-only sheet).
    """
    if raw_df.empty or raw_df.shape[0] < 1 or raw_df.shape[1] < 1:
        return [], None, None

    # ── 1. Header row containing a Question column ─────────────────────
    q_idx = _find_header_row(raw_df)
    if q_idx is not None:
        header_row = raw_df.iloc[q_idx]
        header_cells = _dedupe_columns([
            str(v).strip() if pd.notna(v) and str(v).strip() else f"Unnamed_{i}"
            for i, v in enumerate(header_row)
        ])
        question_col = _match_question_column(header_cells)
        df = raw_df.iloc[q_idx + 1:].reset_index(drop=True)
        df.columns = header_cells
        # _find_header_row already confirmed a question column exists.
        # Defensive guard for the rare case where dedupe rewrites the matched name.
        if question_col is None:
            return [], None, None
        questions = df[question_col].dropna().astype(str).tolist()
        questions = [q.strip() for q in questions if _is_valid_question(q)]
        if not questions:
            return [], str(question_col), f"Sheet \"{sheet_name}\": No valid questions found."
        return questions, str(question_col), None

    # ── 2. Likely header without a Question column ─────────────────────
    h_idx = _find_likely_header_row(raw_df)
    if h_idx is not None:
        data_below = raw_df.iloc[h_idx + 1:]
        if data_below.empty or not _sheet_has_substantial_data(data_below):
            # Summary / index / utility sheet — skip silently.
            return [], None, None
        # Real data with a non-Question header → tell the user.
        header_row = raw_df.iloc[h_idx]
        col_names = [
            str(v).strip() for v in header_row.tolist()
            if pd.notna(v) and str(v).strip() and not str(v).strip().lower().startswith("unnamed")
        ]
        return [], None, (
            f"Sheet \"{sheet_name}\": Could not find a question column. "
            f"Found columns: [{', '.join(col_names)}]. "
            f"Please rename the column containing questions to \"Question\"."
        )

    # ── 3. No header anywhere, but maybe headerless questions present ──
    if not _sheet_has_substantial_data(raw_df):
        return [], None, None  # nothing useful in this sheet

    df = raw_df.copy()
    df.columns = [f"Column_{i+1}" for i in range(df.shape[1])]
    text_cols = list(df.columns)
    if len(text_cols) == 1:
        question_col = text_cols[0]
    else:
        best_col, best_len = text_cols[0], 0
        for col in text_cols:
            avg = df[col].dropna().astype(str).str.len().mean()
            if pd.notna(avg) and avg > best_len:
                best_len, best_col = avg, col
        question_col = best_col
    questions = df[question_col].dropna().astype(str).tolist()
    questions = [q.strip() for q in questions if _is_valid_question(q)]
    if not questions:
        return [], str(question_col), f"Sheet \"{sheet_name}\": No valid questions found."
    return questions, "(no header row — auto-detected)", None


@docprocess_bp.route("/api/docprocess/upload", methods=["POST"])
@jwt_required()
def upload_excel():
    user_id = get_jwt_identity()

    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400

    file = request.files["file"]
    if not file.filename:
        return jsonify({"error": "No file selected"}), 400

    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    if ext not in ("xlsx", "xls"):
        return jsonify({"error": "Only Excel files (.xlsx, .xls) are supported"}), 400

    # --- size check (read into memory first) ---
    file_bytes = file.read()
    if len(file_bytes) > MAX_FILE_SIZE_MB * 1024 * 1024:
        return jsonify({"error": f"File too large. Maximum size is {MAX_FILE_SIZE_MB} MB."}), 400

    safe_name = re.sub(r"[^a-zA-Z0-9_\-.]", "_", file.filename)
    unique_name = f"{uuid.uuid4().hex[:8]}_{safe_name}"
    filepath = os.path.join(UPLOADS_DIR, unique_name)
    with open(filepath, "wb") as f:
        f.write(file_bytes)

    try:
        # --- read ALL sheets without assuming headers ---
        raw_sheets = pd.read_excel(filepath, sheet_name=None, header=None)

        all_questions = []   # list of (sheet_name, question_text)
        detected_columns = {}
        sheet_errors = []

        for sheet_name, raw_df in raw_sheets.items():
            questions, display_col, error = _parse_sheet(sheet_name, raw_df)
            if error:
                sheet_errors.append(error)
                continue
            if not questions:
                continue
            detected_columns[sheet_name] = display_col
            for q in questions:
                all_questions.append((str(sheet_name), q))

        # If ANY sheet had an error, reject the entire upload with details
        if sheet_errors:
            os.remove(filepath)
            return jsonify({
                "error": "Some sheets have problems. Please fix and re-upload.",
                "sheet_errors": sheet_errors,
            }), 400

        if not all_questions:
            os.remove(filepath)
            return jsonify({"error": "No questions found in any sheet. Each sheet must have a column named \"Question\" (or similar), or contain questions without a header row."}), 400

        if len(all_questions) > MAX_QUESTIONS:
            os.remove(filepath)
            return jsonify({"error": f"Too many questions ({len(all_questions)}). Maximum allowed is {MAX_QUESTIONS}."}), 400

        batch_id = str(uuid.uuid4())
        db = SessionLocal()
        try:
            db.execute(text("""
                INSERT INTO batch_jobs (
                    id, user_id, original_filename, original_filepath, status, total_questions, detected_columns, updated_at
                )
                VALUES (:id, :user_id, :filename, :filepath, 'uploaded', :total, :detected_columns, :updated_at)
            """), {
                "id": batch_id,
                "user_id": user_id,
                "filename": file.filename,
                "filepath": filepath,
                "total": len(all_questions),
                "detected_columns": json.dumps(detected_columns),
                "updated_at": datetime.utcnow(),
            })

            # Batch insert all questions at once for speed
            question_rows = [
                {
                    "id": str(uuid.uuid4()),
                    "batch_id": batch_id,
                    "row_index": idx,
                    "question": question,
                    "sheet_name": sheet_name,
                }
                for idx, (sheet_name, question) in enumerate(all_questions)
            ]
            if question_rows:
                db.execute(text("""
                    INSERT INTO batch_questions (id, batch_id, row_index, question, status, sheet_name)
                    VALUES (:id, :batch_id, :row_index, :question, 'pending', :sheet_name)
                """), question_rows)

            db.commit()
            log_audit(user_id, None, "excel_uploaded", "batch_job", batch_id, f"File: {file.filename}, Questions: {len(all_questions)}, Sheets: {len(raw_sheets)}")

            return jsonify({
                "batch_id": batch_id,
                "filename": file.filename,
                "total_questions": len(all_questions),
                "sheets": list(raw_sheets.keys()),
                "detected_columns": detected_columns,
            }), 201
        except Exception as exc:
            db.rollback()
            logger.error("Upload DB error: %s", exc)
            if os.path.exists(filepath):
                os.remove(filepath)
            return jsonify({"error": "Failed to save upload. Please try again."}), 500
        finally:
            db.close()
    except Exception as exc:
        logger.error("Excel parse error: %s", exc)
        if os.path.exists(filepath):
            os.remove(filepath)
        return jsonify({"error": "Could not read the Excel file. Please ensure it is a valid .xlsx file."}), 400


@docprocess_bp.route("/api/docprocess/jobs", methods=["GET"])
@jwt_required()
def list_jobs():
    user_id = get_jwt_identity()
    db = SessionLocal()
    try:
        rows = db.execute(text("""
            SELECT id, original_filename, status, total_questions, processed_count, accepted_count,
                   cached_count, error_count, cancel_requested, created_at, updated_at, started_at, completed_at
            FROM batch_jobs
            WHERE user_id = :user_id
            ORDER BY created_at DESC
        """), {"user_id": user_id}).mappings().all()

        return jsonify([{
            "id": str(row["id"]),
            "filename": row["original_filename"],
            "status": row["status"],
            "total_questions": row["total_questions"],
            "processed_count": row["processed_count"],
            "accepted_count": row["accepted_count"],
            "cached_count": row["cached_count"] or 0,
            "error_count": row["error_count"] or 0,
            "cancel_requested": row["cancel_requested"] or False,
            "created_at": row["created_at"].isoformat() if row["created_at"] else None,
            "updated_at": row["updated_at"].isoformat() if row["updated_at"] else None,
            "started_at": row["started_at"].isoformat() if row["started_at"] else None,
            "completed_at": row["completed_at"].isoformat() if row["completed_at"] else None,
        } for row in rows])
    finally:
        db.close()


@docprocess_bp.route("/api/docprocess/jobs/<batch_id>", methods=["GET"])
@jwt_required()
def get_job(batch_id):
    user_id = get_jwt_identity()
    db = SessionLocal()
    try:
        job = db.execute(text("""
            SELECT id, original_filename, status, total_questions, processed_count, accepted_count,
                   cached_count, error_count, cancel_requested, created_at, started_at, completed_at, detected_columns
            FROM batch_jobs
            WHERE id = :id AND user_id = :user_id
        """), {"id": batch_id, "user_id": user_id}).mappings().fetchone()

        if not job:
            return jsonify({"error": "Job not found"}), 404

        # Pagination: ?page=1&per_page=50 (default: all)
        page = request.args.get("page", type=int)
        per_page = request.args.get("per_page", 50, type=int)
        per_page = min(per_page, 200)  # cap

        if page and page > 0:
            offset = (page - 1) * per_page
            questions = db.execute(text("""
                SELECT id, row_index, question, answer, sources, status, cached, llm_model,
                       retrieval_strategy, latency_ms, error_details, sheet_name
                FROM batch_questions
                WHERE batch_id = :batch_id
                ORDER BY row_index ASC
                LIMIT :limit OFFSET :offset
            """), {"batch_id": batch_id, "limit": per_page, "offset": offset}).mappings().all()
        else:
            questions = db.execute(text("""
                SELECT id, row_index, question, answer, sources, status, cached, llm_model,
                       retrieval_strategy, latency_ms, error_details, sheet_name
                FROM batch_questions
                WHERE batch_id = :batch_id
                ORDER BY row_index ASC
            """), {"batch_id": batch_id}).mappings().all()

        raw_cols = job["detected_columns"]
        if isinstance(raw_cols, str):
            try:
                detected_cols = json.loads(raw_cols)
            except Exception:
                detected_cols = {}
        else:
            detected_cols = raw_cols or {}

        result = {
            "id": str(job["id"]),
            "filename": job["original_filename"],
            "status": job["status"],
            "total_questions": job["total_questions"],
            "processed_count": job["processed_count"],
            "accepted_count": job["accepted_count"],
            "cached_count": job["cached_count"] or 0,
            "error_count": job["error_count"] or 0,
            "cancel_requested": job["cancel_requested"] or False,
            "created_at": job["created_at"].isoformat() if job["created_at"] else None,
            "started_at": job["started_at"].isoformat() if job["started_at"] else None,
            "completed_at": job["completed_at"].isoformat() if job["completed_at"] else None,
            "detected_columns": detected_cols,
            "questions": [{
                "id": str(question["id"]),
                "row_index": question["row_index"],
                "question": question["question"],
                "answer": question["answer"],
                "sources": _serialize_sources(question["sources"]),
                "status": question["status"],
                "cached": question["cached"] or False,
                "llm_model": question["llm_model"],
                "retrieval_strategy": question["retrieval_strategy"],
                "latency_ms": question["latency_ms"],
                "error_details": question["error_details"],
                "sheet_name": question["sheet_name"] or "Sheet1",
            } for question in questions],
        }
        if page and page > 0:
            result["page"] = page
            result["per_page"] = per_page

        return jsonify(result)
    finally:
        db.close()


@docprocess_bp.route("/api/docprocess/jobs/<batch_id>/process", methods=["POST"])
@jwt_required()
def process_job(batch_id):
    from backend.services.rag_service import has_indexed_documents
    if not has_indexed_documents():
        return jsonify({"error": "No knowledge base documents are indexed. Please configure SharePoint sync and index documents in the Admin panel before processing."}), 400

    user_id = get_jwt_identity()
    db = SessionLocal()
    try:
        job = db.execute(text("""
            SELECT id, status
            FROM batch_jobs
            WHERE id = :id AND user_id = :user_id
        """), {"id": batch_id, "user_id": user_id}).mappings().fetchone()
        if not job:
            return jsonify({"error": "Job not found"}), 404

        if job["status"] == "processing":
            # Check if there are actually active questions being processed
            active = db.execute(text("""
                SELECT COUNT(*) FROM batch_questions
                WHERE batch_id = :batch_id AND status IN ('pending', 'processing')
            """), {"batch_id": batch_id}).scalar()
            if active > 0:
                return jsonify({"error": "Job is already being processed"}), 409
            # No active questions — stale status, allow restart
            logger.info("Job %s has status=processing but 0 active questions, allowing restart", batch_id)

        if job["status"] == "canceling":
            # Force-cancel any remaining questions so we can restart cleanly
            db.execute(text("""
                UPDATE batch_questions
                SET status = 'canceled', updated_at = :now
                WHERE batch_id = :batch_id AND status IN ('pending', 'processing')
            """), {"batch_id": batch_id, "now": datetime.utcnow()})
            db.commit()

        db.execute(text("""
            UPDATE batch_jobs
            SET status = 'processing',
                cancel_requested = false,
                started_at = COALESCE(started_at, :now),
                completed_at = NULL,
                updated_at = :now
            WHERE id = :id
        """), {"id": batch_id, "now": datetime.utcnow()})
        db.commit()
    finally:
        db.close()

    # Dispatch questions to Celery task queue
    from backend.tasks import dispatch_batch
    try:
        count = dispatch_batch(batch_id)
    except Exception as exc:
        logger.error("Failed to dispatch batch %s: %s", batch_id, exc)
        # Revert job status so user can retry
        db2 = SessionLocal()
        try:
            db2.execute(text("""
                UPDATE batch_jobs SET status = 'uploaded', cancel_requested = false,
                                     updated_at = :now WHERE id = :id
            """), {"id": batch_id, "now": datetime.utcnow()})
            db2.commit()
        finally:
            db2.close()
        error_msg = str(exc)
        if "No Celery workers" in error_msg:
            return jsonify({"error": "No background workers available. Please start the Celery worker and try again."}), 503
        return jsonify({"error": "Task queue unavailable. Please ensure Redis is running and try again."}), 503
    log_audit(user_id, None, "batch_processing_started", "batch_job", batch_id, f"Started processing {count} questions")
    return jsonify({"message": f"Generation started — {count} questions queued", "batch_id": batch_id})


@docprocess_bp.route("/api/docprocess/jobs/<batch_id>/stop", methods=["POST"])
@jwt_required()
def stop_job(batch_id):
    user_id = get_jwt_identity()
    db = SessionLocal()
    try:
        job = db.execute(text("""
            SELECT id, status
            FROM batch_jobs
            WHERE id = :id AND user_id = :user_id
        """), {"id": batch_id, "user_id": user_id}).mappings().fetchone()
        if not job:
            return jsonify({"error": "Job not found"}), 404

        if job["status"] not in ("processing", "canceling"):
            return jsonify({"error": "Only active jobs can be stopped"}), 409

        db.execute(text("""
            UPDATE batch_jobs
            SET cancel_requested = true,
                status = 'canceling',
                updated_at = :now
            WHERE id = :id
        """), {"id": batch_id, "now": datetime.utcnow()})

        # Only cancel questions still in the queue (pending).
        # Questions already being processed by Celery workers will check
        # cancel_requested themselves and self-cancel after their LLM call.
        db.execute(text("""
            UPDATE batch_questions
            SET status = 'canceled', updated_at = :now
            WHERE batch_id = :batch_id AND status = 'pending'
        """), {"batch_id": batch_id, "now": datetime.utcnow()})

        _refresh_job_counters(db, batch_id)

        # Check if all questions are done (no pending or processing left)
        active = db.execute(text("""
            SELECT COUNT(*) FROM batch_questions
            WHERE batch_id = :batch_id AND status IN ('pending', 'processing')
        """), {"batch_id": batch_id}).scalar()
        if active == 0:
            db.execute(text("""
                UPDATE batch_jobs
                SET status = 'canceled', completed_at = :now, updated_at = :now
                WHERE id = :id AND status = 'canceling'
            """), {"id": batch_id, "now": datetime.utcnow()})

        db.commit()
    finally:
        db.close()

    log_audit(user_id, None, "batch_processing_stop_requested", "batch_job", batch_id, "User requested to stop batch processing")
    return jsonify({"message": "Stop requested", "batch_id": batch_id})


@docprocess_bp.route("/api/docprocess/questions/<question_id>/regenerate", methods=["POST"])
@jwt_required()
def regenerate_answer(question_id):
    user_id = get_jwt_identity()
    db = SessionLocal()
    try:
        row = db.execute(text("""
            SELECT bq.id, bq.question, bq.batch_id
            FROM batch_questions bq
            JOIN batch_jobs bj ON bq.batch_id = bj.id
            WHERE bq.id = :id AND bj.user_id = :user_id
        """), {"id": question_id, "user_id": user_id}).mappings().fetchone()

        if not row:
            return jsonify({"error": "Question not found"}), 404

        started = time.perf_counter()
        result = answer_question_with_sources(row["question"], use_cache=False)
        latency_ms = int((time.perf_counter() - started) * 1000)

        db.execute(text("""
            UPDATE batch_questions
            SET answer = :answer,
                sources = :sources,
                status = 'answered',
                cached = false,
                llm_model = :llm_model,
                retrieval_strategy = :retrieval_strategy,
                latency_ms = :latency_ms,
                error_details = NULL,
                updated_at = :now
            WHERE id = :id
        """), {
            "id": question_id,
            "answer": result["answer"],
            "sources": json.dumps(result["sources"]),
            "llm_model": result.get("model"),
            "retrieval_strategy": result.get("retrieval_strategy"),
            "latency_ms": latency_ms,
            "now": datetime.utcnow(),
        })
        _refresh_job_counters(db, str(row["batch_id"]))
        db.commit()

        return jsonify({
            "id": question_id,
            "answer": result["answer"],
            "sources": result["sources"],
            "status": "answered",
            "latency_ms": latency_ms,
            "retrieval_strategy": result.get("retrieval_strategy"),
        })
    finally:
        db.close()


@docprocess_bp.route("/api/docprocess/questions/<question_id>/edit", methods=["PUT"])
@jwt_required()
def edit_answer(question_id):
    user_id = get_jwt_identity()
    data = request.get_json() or {}
    new_answer = data.get("answer", "").strip()
    if not new_answer:
        return jsonify({"error": "Answer cannot be empty"}), 400

    db = SessionLocal()
    try:
        row = db.execute(text("""
            SELECT bq.batch_id
            FROM batch_questions bq
            JOIN batch_jobs bj ON bq.batch_id = bj.id
            WHERE bq.id = :id AND bj.user_id = :user_id
        """), {"id": question_id, "user_id": user_id}).mappings().fetchone()

        if not row:
            return jsonify({"error": "Question not found"}), 404

        db.execute(text("""
            UPDATE batch_questions
            SET answer = :answer,
                status = 'answered',
                cached = false,
                updated_at = :now
            WHERE id = :id
        """), {"id": question_id, "answer": new_answer, "now": datetime.utcnow()})
        _refresh_job_counters(db, str(row["batch_id"]))
        db.commit()
        return jsonify({"message": "Answer updated"})
    finally:
        db.close()


@docprocess_bp.route("/api/docprocess/questions/<question_id>/accept", methods=["POST"])
@jwt_required()
def accept_answer(question_id):
    user_id = get_jwt_identity()
    db = SessionLocal()
    try:
        row = db.execute(text("""
            SELECT bq.batch_id
            FROM batch_questions bq
            JOIN batch_jobs bj ON bq.batch_id = bj.id
            WHERE bq.id = :id AND bj.user_id = :user_id
        """), {"id": question_id, "user_id": user_id}).mappings().fetchone()
        if not row:
            return jsonify({"error": "Question not found"}), 404

        db.execute(text("""
            UPDATE batch_questions
            SET status = 'accepted', updated_at = :now
            WHERE id = :id
        """), {"id": question_id, "now": datetime.utcnow()})
        _refresh_job_counters(db, str(row["batch_id"]))
        db.commit()
        return jsonify({"message": "Answer accepted"})
    finally:
        db.close()


@docprocess_bp.route("/api/docprocess/jobs/<batch_id>/accept-all", methods=["POST"])
@jwt_required()
def accept_all(batch_id):
    user_id = get_jwt_identity()
    db = SessionLocal()
    try:
        job = db.execute(text("""
            SELECT id
            FROM batch_jobs
            WHERE id = :id AND user_id = :user_id
        """), {"id": batch_id, "user_id": user_id}).fetchone()
        if not job:
            return jsonify({"error": "Job not found"}), 404

        db.execute(text("""
            UPDATE batch_questions
            SET status = 'accepted', updated_at = :now
            WHERE batch_id = :batch_id AND status = 'answered'
        """), {"batch_id": batch_id, "now": datetime.utcnow()})
        stats = _refresh_job_counters(db, batch_id)
        db.commit()
        return jsonify({"message": f"{stats['accepted_count']} answers accepted", "accepted_count": stats["accepted_count"]})
    finally:
        db.close()


@docprocess_bp.route("/api/docprocess/jobs/<batch_id>/download", methods=["GET"])
@jwt_required()
def download_results(batch_id):
    user_id = get_jwt_identity()
    db = SessionLocal()
    try:
        job = db.execute(text("""
            SELECT id, original_filename
            FROM batch_jobs
            WHERE id = :id AND user_id = :user_id
        """), {"id": batch_id, "user_id": user_id}).mappings().fetchone()
        if not job:
            return jsonify({"error": "Job not found"}), 404

        questions = db.execute(text("""
            SELECT question, answer, sources, status, latency_ms, llm_model, cached, retrieval_strategy, sheet_name
            FROM batch_questions
            WHERE batch_id = :batch_id
            ORDER BY row_index ASC
        """), {"batch_id": batch_id}).fetchall()

        # Group questions by sheet
        from collections import OrderedDict
        sheets_data = OrderedDict()
        for q in questions:
            sheet = q[8] or "Sheet1"
            sheets_data.setdefault(sheet, []).append(q)

        wb = openpyxl.Workbook()
        # Remove default sheet — we'll create named ones
        wb.remove(wb.active)

        header_font = openpyxl.styles.Font(bold=True, size=11, color="FFFFFF")
        header_fill = openpyxl.styles.PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        header_align = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)
        wrap_align = openpyxl.styles.Alignment(wrap_text=True, vertical="top")

        headers = ["#", "Question", "Answer", "Status", "Sources"]

        status_map = {
            'answered': 'Accepted',
            'accepted': 'Accepted',
            'processing': 'Not Reviewed',
            'pending': 'Not Reviewed',
            'canceled': 'Not Reviewed',
            'error': 'Error',
        }

        def _clean_markdown(text):
            if not text:
                return ""
            import re as _re
            # Convert ### headings to plain text with newline
            text = _re.sub(r'^#{1,4}\s+', '', text, flags=_re.MULTILINE)
            # Convert - bullet to bullet char
            text = _re.sub(r'^\s*[-*]\s+', '\u2022 ', text, flags=_re.MULTILINE)
            return text.strip()

        def _write_rich_answer(ws, row, col, text, wrap_align):
            """Write answer with bold formatting preserved via openpyxl rich text."""
            import re as _re
            cleaned = _clean_markdown(text)
            if '**' not in cleaned:
                cell = ws.cell(row=row, column=col, value=cleaned)
                cell.alignment = wrap_align
                return

            from openpyxl.cell.rich_text import CellRichText, TextBlock
            from openpyxl.cell.text import InlineFont
            parts = _re.split(r'(\*\*.*?\*\*)', cleaned)
            rich_parts = []
            bold_font = InlineFont(b=True)
            for part in parts:
                if part.startswith('**') and part.endswith('**'):
                    rich_parts.append(TextBlock(bold_font, part[2:-2]))
                elif part:
                    rich_parts.append(part)
            if rich_parts:
                ws.cell(row=row, column=col).value = CellRichText(*rich_parts)
            else:
                ws.cell(row=row, column=col, value=cleaned)
            ws.cell(row=row, column=col).alignment = wrap_align

        for sheet_name, sheet_questions in sheets_data.items():
            # Truncate sheet name to Excel's 31-char limit
            ws = wb.create_sheet(title=sheet_name[:31])

            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align

            for row_idx, question in enumerate(sheet_questions, 2):
                sources_list = _serialize_sources(question[2])
                source_text = "\n".join(
                    f"{source.get('document', '')} - {source.get('section', '')}".strip(" -")
                    for source in sources_list
                )
                raw_status = question[3] or "pending"

                ws.cell(row=row_idx, column=1, value=row_idx - 1)
                cell_q = ws.cell(row=row_idx, column=2, value=question[0])
                cell_q.alignment = wrap_align
                _write_rich_answer(ws, row_idx, 3, question[1] or "", wrap_align)
                cell_s = ws.cell(row=row_idx, column=4, value=status_map.get(raw_status, raw_status))
                cell_s.alignment = wrap_align
                cell_src = ws.cell(row=row_idx, column=5, value=source_text)
                cell_src.alignment = wrap_align

            for name, width in {"A": 6, "B": 50, "C": 70, "D": 16, "E": 40}.items():
                ws.column_dimensions[name].width = width

        safe_name = re.sub(r"[^a-zA-Z0-9_\-.]", "_", job["original_filename"])
        result_filename = f"results_{safe_name}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        result_path = os.path.realpath(os.path.join(RESULTS_DIR, result_filename))
        if not result_path.startswith(os.path.realpath(RESULTS_DIR)):
            return jsonify({"error": "Invalid filename"}), 400

        wb.save(result_path)
        file_size = os.path.getsize(result_path)
        download_id = str(uuid.uuid4())

        db.execute(text("""
            INSERT INTO downloads (id, user_id, filename, file_type, file_path, file_size)
            VALUES (:id, :user_id, :filename, 'batch_result', :file_path, :file_size)
        """), {
            "id": download_id,
            "user_id": user_id,
            "filename": result_filename,
            "file_path": result_path,
            "file_size": file_size,
        })
        db.commit()

        log_audit(user_id, None, "batch_downloaded", "batch_job", batch_id, f"Downloaded: {result_filename}")
        return send_file(result_path, as_attachment=True, download_name=result_filename)
    except Exception as exc:
        logger.error("Download error: %s", exc)
        return jsonify({"error": "Failed to generate download."}), 500
    finally:
        db.close()


@docprocess_bp.route("/api/docprocess/jobs/<batch_id>", methods=["DELETE"])
@jwt_required()
def delete_job(batch_id):
    user_id = get_jwt_identity()
    db = SessionLocal()
    try:
        job = db.execute(text("""
            SELECT id, original_filepath
            FROM batch_jobs
            WHERE id = :id AND user_id = :user_id
        """), {"id": batch_id, "user_id": user_id}).mappings().fetchone()
        if not job:
            return jsonify({"error": "Job not found"}), 404

        if os.path.exists(job["original_filepath"]):
            os.remove(job["original_filepath"])

        db.execute(text("DELETE FROM batch_jobs WHERE id = :id"), {"id": batch_id})
        db.commit()
        return jsonify({"message": "Job deleted"})
    except Exception as exc:
        db.rollback()
        logger.error("Delete job error: %s", exc)
        return jsonify({"error": "Failed to delete job"}), 500
    finally:
        db.close()
